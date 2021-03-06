# -*- coding: utf-8 -*-
"""
Created on Fri Jul 21 23:22:44 2017

@author: HX301
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jul 20 12:10:44 2017

@author: ADBC
"""
#from os import getcwd,makedirs
import os
import sys
import string
import ConfigParser
import codecs 
import tushare as ts
import pandas as pd
import win32com.client as win32
from urllib import urlopen
from HTMLParser import HTMLParser
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import BLUE

#Init
#http://money.finance.sina.com.cn/corp/go.php/vDOWN_BalanceSheet/displaytype/4/stockid/002450/ctrl/all.phtml
#http://money.finance.sina.com.cn/corp/go.php/vDOWN_ProfitStatement/displaytype/4/stockid/002450/ctrl/all.phtml
#http://money.finance.sina.com.cn/corp/go.php/vDOWN_CashFlow/displaytype/4/stockid/002450/ctrl/all.phtml

def log(msg, level=1):  
    '''''log message''' 
    if (level < 3):
        print msg  
   

#合并某stock在SINA三个报表文件orifile到mergefile,当first为0时读表头（时间）
#如需要建立3个sheet，只需要将first改成sheetnum就可以了
def excelMerge(stock, mergefile, orifile, first):
    if not mergefile or not orifile:
        log("no file to merge or do not know mergefile", 3)
        return
    
    #创建新的Excel 2007+文件
    if (first == 0):
        workbook = Workbook()
        worksheet = workbook.worksheets[0]    
        worksheet.title = stock
        #workbook.create_sheet(title='300072',index=0)
    else:
        workbook = load_workbook(mergefile)
        worksheet = workbook.worksheets[0]    
    
    data = load_workbook(orifile)
    for sheetnum, sheet in enumerate(data.worksheets):
        #根据设定的表头行数，设置读取的起始行
        #第一个sheet读取表头，后面的sheet忽略表头
        if first == 0:
            rowStart = 0
        else:
            rowStart = 2 #跨过财报表头
        #遍历原sheet，根据情况忽略表头
        rlist = list(sheet.rows)
        for row in rlist[rowStart:]:
            line = [col.value for col in row]
            worksheet.append(line)
    #保存新文件
    workbook.save(mergefile)

# Get url netfile and write it into fname
def getNetFile(url, fname):
    sock = urlopen(url)
    htmlCode = sock.read()
    sock.close
    fp = open(fname,"wb")
    fp.seek(0)
    fp.truncate()
    fp.write(htmlCode)
    fp.close
    
# change xls file to xlsx file
def chgExcelFmt(fname):
    excel = win32.Dispatch('Excel.Application')
    ori_alert_status = excel.DisplayAlerts
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(Filename=fname)
        wb.SaveAs(fname+"x", FileFormat = 51, Local=True)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
    finally:
        excel.DisplayAlerts = ori_alert_status
        excel.Quit()
        del excel
        
'''获取新浪三报表并合并成一个报表'''   
def getSinaExcel(stock, stockname, savepath):
    b_asset = 'http://money.finance.sina.com.cn/corp/go.php/vDOWN_BalanceSheet/displaytype/4/stockid/'
    b_profit = 'http://money.finance.sina.com.cn/corp/go.php/vDOWN_ProfitStatement/displaytype/4/stockid/'
    b_cash = 'http://money.finance.sina.com.cn/corp/go.php/vDOWN_CashFlow/displaytype/4/stockid/'
    lasturl = '/ctrl/all.phtml'
    
    mergefile = savepath+'\\'+stockname+'.xlsx'
    tmpfile =  savepath+'\\'+'tmp.xls'
    
    downurl = b_asset+stock+lasturl
    getNetFile(downurl, tmpfile)
    chgExcelFmt(tmpfile)
    excelMerge(stock, mergefile, tmpfile+'x', 0)

    downurl = b_profit+stock+lasturl
    getNetFile(downurl, tmpfile)
    chgExcelFmt(tmpfile)
    excelMerge(stock, mergefile, tmpfile+'x', 1)
       
    downurl = b_cash+stock+lasturl
    getNetFile(downurl, tmpfile)
    chgExcelFmt(tmpfile)
    excelMerge(stock, mergefile, tmpfile+'x', 2)

def dealExcel1(fname):
    wb = load_workbook(fname)
    ws = wb.worksheets[0]  
    #df = openpyxl.utils.dataframe(ws.values)    
    df = pd.DataFrame(ws.values) 
    df.set_index(df[0])
    df.columns = df.loc[0]
    
    print df.loc[0]
    #print df['单位']
    
    #df.set_index = 
    #df.to_excel('foo.xlsx', sheet_name='sheet1')
    print df.index, df.columns

def gbk2utf(in_data , tag):  
    if 1 == tag:  
        return in_data.encode('gb2312').decode('gb2312')  
    elif 0 == tag:  
        return in_data.encode('gb2312').decode('gb2312').encode('utf-8')  
        
def getExcelDF(fname):
    wb = load_workbook(fname)
    ws = wb.worksheets[0]  

    data = ws.values
    cols = next(data)[1:]   #时间
    data = list(data)
    #idx = [gbk2utf(r[0], 0) for r in data]
    idx = [r[0] for r in data]  #科目
    #print type(cols), cols
    #print type(cols[0]), cols[0]
    #print '\\n'
    #print type(idx), idx[3]
    data = (r[1:] for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)

    #print df.index, df.columns
    #print df[20170630][u'一、营业收入']
    #string = '一、营业收入'
    #print df[20170630][string.decode('utf-8')]
    return df

#环比
def qoqResult():
    lst = []
    return lst

#同比
def yoyResult(df, cols, cmpItem):
    log(u'报表比较对象： '+ cmpItem, 3)
    lst = [-1 for i in cols]
    cur = cols[0]
    size = len(cols)-1 #get rid of the last 'None' column
    n = 0
    for cur in cols:
        if (n+5 < size):
            lst1 = cols[n+1:n+5]
        else:
            lst1 = cols[n+1:size]
        #print lst1 
        for i in lst1:
            if i and (cur - i == 10000):
                #print cur, i
                if df[i][cmpItem]:  #Not null\0\''
                    ration = (df[cur][cmpItem] - df[i][cmpItem])/abs(float(df[i][cmpItem]))
                else:
                    ration = 1000  #假定一个极限增长率
                #print df[cur][cmpItem], df[i][cmpItem], ration
                lst[n] = round(ration,3)*100    #百分比
        n = n + 1
    log(lst, 3)
    return lst

#In: 已经生成的原始EXCEL报表；
#分析
def DFAnalyse(df):
    global PARSE_CALC_LIST
    global PARSE_KEY_LIST
    
    #newIdx = dics.keys()+PARSE_KEY_LIST
    newIdx = PARSE_KEY_LIST
    #df = df.dropna(axis=1, how='any')
    cols = df.columns

    df_calc = pd.DataFrame(index=newIdx, columns=cols)
    # 同比增长率
    '''
    for idx in dics.keys():
        lstData = yoyResult(df,cols, dics[idx])
        df_calc.loc[idx] = lstData
        #print newdf
    '''
    i = 0
    calc_list = PARSE_CALC_LIST
    while (i < len(calc_list)):
        if (str(calc_list[i+2]).find('^10') > -1): #同比增长率[0]
            result_list = yoyResult(df,cols, calc_list[i+1])
        elif (str(calc_list[i+2]).find('^20') > -1): #环比增长率
            result_list = [-1 for k in cols]
        else:
            col_n = 0
            result_list = [-1 for k in cols]
            for col in df.columns:
                if (col is None):
                    continue
                j = 0
                result = ''
                while (j < len(calc_list[i+1])):
                    if (j == 0):
                        result = df[col][calc_list[i+1][j]]
                        if (result is None or result != result): #None or Non in excel cell
                            break
                    else:
                        var = df[col][calc_list[i+1][j]]
                        if (var is None or var != var or (calc_list[i+2][j-1]=='/' and var==0)):
                            break
                        tmp = str(result)+' '+ calc_list[i+2][j-1] + ' '+str(var)
                        log(tmp, 3)
                        result = eval(tmp) #没有考虑百分比为负的情况
                        log(result, 3)
                    j += 1
                #改变格式后进行记录
                if (result and abs(result) > 10000000): #1千万以上，一般为金额
                   result = round(result/100000000, 1) #将数字万单位转为亿
                elif (result and result < 1000): #一般为百分比
                   result = round(result, 3)*100
                result_list[col_n] = result
                col_n += 1
            
        df_calc.loc[calc_list[i]] = result_list
        i += 3
        
    return df_calc

"""
暂未用
padas dataframe生成excel
"""
def dataFrame2sheet(dataframe,excelWriter):

   # DataFrame转换成excel中的sheet表
   dataframe.to_excel(excel_writer=excelWriter, sheet_name="info1",index=None)
   dataframe.to_excel(excel_writer=excelWriter, sheet_name="info2",index=None)
   dataframe.to_excel(excel_writer=excelWriter, sheet_name="info3",index=None)

   excelWriter.save()
   excelWriter.close()
   
"""
excel中新增1个sheet表,如存在则先删除sheet后再重写
"""
def excelAddSheet(dataframe,excelWriter,sheetname):

   book = load_workbook(excelWriter.path)
   excelWriter.book = book
   sheets = book.get_sheet_names()
   if (sheetname in sheets):
       book.remove_sheet(book.get_sheet_by_name(sheetname))
   dataframe.to_excel(excel_writer=excelWriter,sheet_name=sheetname)
   excelWriter.close()
   
def dealPath(pathname=''):  
    '''''deal with windows file path'''  
    if pathname:  
        pathname = pathname.strip()  
    if pathname:  
        pathname = r'%s'%pathname  
        pathname = string.replace(pathname, r'/', '\\')  
        pathname = os.path.abspath(pathname)  
        if pathname.find(":\\") == -1:  
            pathname = os.path.join(os.getcwd(), pathname)  
    return pathname  
          
def loadConfig(configfile='./sconfig.ini'):  
    '''''parse config file'''  
    global SPATH, DPATH  
    global STOCK_LIST, DOWN_LIST, REPORT_LIST
    global PARSE_CALC_LIST, PARSE_KEY_LIST
      
    file = dealPath(configfile)  
    if not os.path.isfile(file):  
        log('Can not find the config.ini',3)  
        return False  
    parser = ConfigParser.ConfigParser()  
    #parser.read(file)  
    parser.readfp(codecs.open(file, "r", "utf-8-sig"))
    SPATH = parser.get('pathconfig', 'spath').strip()  
    DPATH = parser.get('pathconfig', 'dpath').strip()  
    stocklist = parser.get('stockconfig', 'stocklist').strip()  
    downlist = parser.get('stockconfig', 'stockdown').strip()
    reportlist = parser.get('stockconfig', 'stockreport').strip()
    if stocklist:  
        STOCK_LIST = stocklist.split(";")  
    if downlist: 
        print 'get down?'
        DOWN_LIST = downlist.split(";")  
    if reportlist:  
        REPORT_LIST = reportlist.split(";")  
    #parse stock analyze
    calckeylist = parser.options('calcuconfig')
    PARSE_KEY_LIST = calckeylist
    PARSE_CALC_LIST = [-1 for i in calckeylist*3] #3 item in 1 key
    i = 0
    for key in calckeylist:  #如果excel顺序不对,打印或修改这条语句
        PARSE_CALC_LIST[i] = key
        val = parser.get('calcuconfig', key).strip() 
        vallist = val.split(':')
        if (vallist[0].find(',') == -1 and vallist[1].find(',') == -1): #同比或环比
            PARSE_CALC_LIST[i+1] = vallist[0]
            PARSE_CALC_LIST[i+2] = vallist[1]
        else:
            PARSE_CALC_LIST[i+1] = vallist[0].split(',')
            PARSE_CALC_LIST[i+2] = vallist[1].split(',')
        i += 3
    
def is_number(s):
    try:
        if (s and s != ''):
            float(s)
            return True
    except ValueError:
        pass
    return False

def len_cell(s):
    try:
        if (s and s != ''):
            l = len(str(s))
            return l
        else:
            return 0
    except (TypeError, ValueError):
        pass
    return len('u\''+s) #中文

#s: Long, null or 20171231...
def is_year(l):
    s = ''
    try:
        if (l):
            s = str(l)
        if (s == ''):
            return False
        elif (s.endswith('1231')):  #ex:20171231
            return True
        else:
            return False
    except (TypeError, ValueError): #str为中文报错
        return False

class stnumParser(HTMLParser):

    def __init__(self):
        HTMLParser.__init__(self)
        self._count = 0
        self._events = dict()
        self._flag = None
        self._item = None

    def handle_starttag(self, tag, attrs):
        if tag == 'div' and attrs.__contains__(('align', 'center')) and not (attrs.__contains__(('class', 'text'))):
            self._count += 1
            #self._events[self._count] = dict()
            self._flag = 'stnum-change'

    def handle_data(self, data):
        if self._flag == 'stnum-change':
           #print data, type(data), self._count
           if (self._count % 2 == 0):
               self._events[self._item] = data #''.join(re.findall(r"\d+\.?\d*", data))
           else:
               self._item = data
        self._flag = None

    def event_list(self):
        print '股本变化次数有：', self._count/2, '次，具体如下：'
        for key in self._events.keys():
            #print type(key)
            print key+ ' : '+ self._events[key]

#获取股本网页，分析并最终得到dicts 时间：股本数
def parseStnum(url):
    try:
        parser = stnumParser()
        sock = urlopen(url)
        htmlCode = sock.read()
        htmlCode = htmlCode.decode("utf8","ignore")
        #print htmlCode
        #sock.close
    except IOError,e:
        print 'IOError:', e
    else:
        parser.feed(htmlCode)
        #parser.event_list()
        #print parser._events
    finally:
        sock.close        
        return parser._events 

#查找离参数日期最近的股本数
def getStocknum(date, sn_dics):
    stnum = ''
    dmin = 0
    dkey = ''
    #print "finding date = ", date
    for key in sn_dics.keys():
        dicdate = key.replace('-', '')
        dnear = date - long(dicdate)
        if (dnear > 0 and dmin == 0):
            dmin = dnear
            dkey = key
        elif (dnear > 0 and dnear < dmin):
            dmin = dnear
            dkey = key
        #print dnear, dmin
        #print 'dkey = '+ dkey
        if (dkey != ''):
            stnum = sn_dics[dkey]
        else: #股本为最小值
            stnum = sn_dics[min(sn_dics, key = sn_dics.get)]
            
    #print "find key = "+ dkey +"and  stnum = " + stnum
    return float(stnum)*10000
    
def getMktval(date, sn_dics, df):
    mktval = 0.0
    #get the date close price
    ndate = str(date)
    ndate = ndate[0:4]+ '-' + ndate[4:6]
    #print ndate
    #假定df内对应ndate的只有一行（月）的收盘价
    close = df[df['date'].str.contains(ndate)].close
    if (close.empty): #还没有上市，无收盘价格
        close = 0
    close = float(close)
    #print type(close), close
    #get the date stock num
    #print date
    stnum = getStocknum(date, sn_dics)
    mktval = close*stnum           
    return mktval
    

STOCK_LIST = []
DOWN_LIST = []
REPORT_LIST = []
PARSE_KEY_LIST = []
PARSE_CALC_LIST = []
WS_DISP = "exdisply"
WS_RESULT = "exresult"
CL_LOWGREEN = "53868B"
          
def main():  
    global STOCK_LIST, DOWN_LIST, REPORT_LIST
    global PARSE_CALC_LIST
    
    defaultencoding = 'utf-8'
    if sys.getdefaultencoding() != defaultencoding:
        reload(sys)
    sys.setdefaultencoding(defaultencoding)
    
    '''''main function  '''
    loadConfig()  
    stockdic = {}
    df_stocks = ts.get_stock_basics()
    #print SPATH, DPATH, MAX_SHEET_INDEX   # del later
    filepath = os.getcwd() + '\\baobiao'
    
    if STOCK_LIST:
        map(lambda x:stockdic.setdefault(x.split(':')[0], x.split(':')[1]), STOCK_LIST)
        #for d,x in stockdic.items():
        #    print d,x
            
    if DOWN_LIST:
        for stock in DOWN_LIST:
            if stock in stockdic.keys():
                #stockname = u'中国平安'
                stockname = stockdic[stock]
            else:
                stockname = df_stocks['name'][stock]
                stockname = unicode(stockname, 'utf-8')
                stockdic.setdefault(stock, stockname)
            log("download "+stockname+ " now ...", 2)
            getSinaExcel(stock, stockname, filepath)
            fname = filepath+'\\'+stockname+'.xlsx'
           
            # SINA合并三大表格后，增加一行总市值数据
            mktvalurl = 'http://money.finance.sina.com.cn/corp/go.php/vCI_StockStructureHistory/stockid/'+stock+'/stocktype/TotalStock.html' 
            sn_dics = parseStnum(mktvalurl)
            #print sn_dics
            df_mprice = ts.get_k_data(stock, ktype='M', autype='None',)
            #日期转换成sina格式（对应月份），20171231	20170930	20170630	20170331
            wb = load_workbook(fname)
            s_sheet = wb.get_sheet_by_name(stock)
            rows_len = len(list(s_sheet.rows))
            cols_len = len(list(s_sheet.columns))
            for col in range(1, cols_len+1):
                if col == 1:
                    s_sheet.cell(row=rows_len+1, column=col).value = '总市值'
                else:
                    date = s_sheet.cell(row=1, column=col).value
                    if (date):
                        mktval = getMktval(date, sn_dics, df_mprice)
                        s_sheet.cell(row=rows_len+1, column=col).value = mktval
            rows_len = rows_len + 1
            
            # EXCEL格式处理:将数字万单位转为亿        
            t_sheet = wb.create_sheet(title=WS_DISP)
            s_sheet.freeze_panes = 'B2'            
            t_sheet.freeze_panes = 'B2'
            #log("sheet name: " +sheet.title, 1)
            dims = {}
            
            #生成易读报表
            #负债和所有者权益(或股东权益)总计  一、营业总收入
            
            row_allasset = row_allgain = 0
            for row in range(1, rows_len+1):
                if s_sheet.cell(row=row, column=1).value == u'负债和所有者权益(或股东权益)总计':
                    row_allasset = row
                if s_sheet.cell(row=row, column=1).value == u'一、营业总收入':
                    row_allgain = row
            #print row_allasset, row_allgain
            ration = 0
            for row in range(1, rows_len+1):
                excol = 0
                for col in range(1, cols_len+1):
                    s_col = col
                    cell_val = s_sheet.cell(row=row, column=col).value
                    if col <> 1:
                        col += excol
                        excol += 1
                    if (row <> 1 and col <> 1 and is_number(cell_val)):
                        #value format: col
                        t_sheet.cell(row=row, column=col).value = cell_val/100000000 #将数字万单位转为亿 
                        if (is_year(t_sheet.cell(row=1, column=col).value)):
                            t_sheet.cell(row=row, column=col).font = Font(color=BLUE, italic=False)    
                        #ration: col+1
                        if (row <= row_allasset) and s_sheet.cell(row=row_allasset, column=s_col).value:
                            ration = cell_val/s_sheet.cell(row=row_allasset, column=s_col).value
                        if (row >= row_allgain) and s_sheet.cell(row=row_allgain, column=s_col).value:
                            ration = cell_val/s_sheet.cell(row=row_allgain, column=s_col).value
                        t_sheet.cell(row=row, column=col+1).value = round(ration, 4)*100
                        t_sheet.cell(row=row, column=col+1).font = Font(color=CL_LOWGREEN, italic=True)
                    else:
                        t_sheet.cell(row=row, column=col).value = cell_val
                        t_sheet.cell(row=row, column=col+1).value = ''
                    #调整列宽度
                    dims[t_sheet.cell(row=row, column=col).column] = max((dims.get(t_sheet.cell(row=row, column=col).column, 0), len_cell(cell_val)))
                    dims[t_sheet.cell(row=row, column=col+1).column] = max((dims.get(t_sheet.cell(row=row, column=col+1).column, 0), len_cell(t_sheet.cell(row=row, column=col+1).value)))
                    for col, value in dims.items():
                        t_sheet.column_dimensions[col].width = value
                    #t_sheet.column_dimensions[col+1].width = 6 #max:100.00
            wb.save(fname)
            #wb.close()

    if REPORT_LIST:
        for stock in REPORT_LIST:
            if (stock not in stockdic.keys()):
                log("exception: find stock: "+stock+" not in dic", 1)
                stockname = df_stocks['name'][stock]
                stockname = unicode(stockname, 'utf-8')
            else:
                stockname = stockdic[stock]
            fname = filepath+'\\'+stockname+'.xlsx'
            if ( not os.access(fname, os.W_OK)):
                log("file: "+ fname + " not exist in directory, download first pls!")
                continue
            excelDF = getExcelDF(fname)
            #log(excelDF.head(), 1)
            log("Report handling: "+stockname, 2)
            newdf = DFAnalyse(excelDF)
            #print newdf
            excelWriter=pd.ExcelWriter(fname,engine='openpyxl')
            excelAddSheet(newdf,excelWriter, WS_RESULT)
            #newdf.to_excel(fname,sheet_name='exresult')
            #dfToNewExcel()
            # EXCEL格式处理:固定宽度  
            wb = load_workbook(fname)
            sheet = wb.get_sheet_by_name(WS_RESULT)
            sheet.freeze_panes = 'B2'
            dims = {}
            rows_len = len(list(sheet.rows))
            cols_len = len(list(sheet.columns))
            for row in range(1, rows_len+1):
                for col in range(1, cols_len+1):
                    if (is_year(sheet.cell(row=1, column=col).value)):
                        sheet.cell(row=row, column=col).font = Font(color=BLUE, italic=False)
            wb.save(fname)
    #raw_input("Please press any key to exit!")  
          
          
if __name__=="__main__":  
    main()  

'''
config.ini(UTF-8编码）文件如下：
[plain] view plain copy

    [pathconfig]  
    #;spath表示需要处理的excel文件目录  
    spath=./tests  
    #;dpath表示处理后的excel文件目录  
    dpath=./dest  
      
    [otherconfig]  
    #;filelist表示不需要做特殊处理的excel文件列表,以英文分号分隔  
    filelist=  
    #;maxindex表示需要处理每个excel文件的前几张表  
    maxindex=1  
    #;deleterows表示需要删除的阿拉伯数字行号，用英文分号分隔  
    deleterows=2;3  
    
    content = open('BaseConfig.cfg').read()  
    #Window下用记事本打开配置文件并修改保存后，编码为UNICODE或UTF-8的文件的文件头  
    #会被相应的加上\xff\xfe（\xff\xfe）或\xef\xbb\xbf，然后再传递给ConfigParser解析的时候会出错  
    #，因此解析之前，先替换掉  
    content = re.sub(r"\xfe\xff","", content)  
    content = re.sub(r"\xff\xfe","", content)  
    content = re.sub(r"\xef\xbb\xbf","", content)  
    open('BaseConfig.cfg', 'w').write(content)  
'''
