# -*- coding: utf-8 -*-
"""
Created on Wed Aug 29 09:51:22 2018

@author: ADBC
"""

import urllib
import tushare as ts
import pandas as pd
from openpyxl import load_workbook
from HTMLParser import HTMLParser
#import sys
#reload(sys)
#sys.setdefaultencoding('GB2312')

def getExcelDF(fname):
    wb = load_workbook(fname)
    ws = wb.worksheets[0]  

    data = ws.values
    cols = next(data)[1:]   #时间
    data = list(data)
    #idx = [gbk2utf(r[0], 0) for r in data]
    idx = [r[0] for r in data]  #科目
    #print type(cols), cols
    #print type(cols[0]), cols[0]
    #print '\\n'
    #print type(idx), idx[3]
    data = (r[1:] for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)

    #print df.index, df.columns
    #print df[20170630][u'一、营业收入']
    #string = '一、营业收入'
    #print df[20170630][string.decode('utf-8')]
    return df

class stnumParser(HTMLParser):

    def __init__(self):
        HTMLParser.__init__(self)
        self._count = 0
        self._events = dict()
        self._flag = None
        self._item = None

    def handle_starttag(self, tag, attrs):
        if tag == 'div' and attrs.__contains__(('align', 'center')) and not (attrs.__contains__(('class', 'text'))):
            self._count += 1
            #self._events[self._count] = dict()
            self._flag = 'stnum-change'

    def handle_data(self, data):
        if self._flag == 'stnum-change':
           print data, type(data), self._count
           if (self._count % 2 == 0):
               self._events[self._item] = data #''.join(re.findall(r"\d+\.?\d*", data))
           else:
               self._item = data
        self._flag = None

    def event_list(self):
        print '股本变化次数有：', self._count/2, '次，具体如下：'
        for key in self._events.keys():
            #print type(key)
            print key+ ' : '+ self._events[key]

''' correct code         
def getNetFile(url, fname):
    sock = urllib.urlopen(url)
    htmlCode = sock.read()
    sock.close
    fp = open(fname,"wb")
    fp.seek(0)
    fp.truncate()
    fp.write(htmlCode)
    fp.close

def getNetHtml(url):
    sock = urllib.urlopen(url)
    htmlCode = sock.read()
    print htmlCode
    sock.close
'''

def parseStnum(url):
    try:
        parser = stnumParser()
        sock = urllib.urlopen(url)
        htmlCode = sock.read()
        htmlCode = htmlCode.decode("utf8","ignore")
        #print htmlCode
        #sock.close
    except IOError,e:
        print 'IOError:', e
    else:
        parser.feed(htmlCode)
        parser.event_list()
        #print parser._events
    finally:
        sock.close        
        return parser._events 

#查找离参数日期最近的股本数
def getStocknum(date, sn_dics):
    stnum = ''
    dmin = 0
    dkey = ''
    print "finding date = ", date
    for key in sn_dics.keys():
        dicdate = key.replace('-', '')
        dnear = date - long(dicdate)
        if (dnear > 0 and dmin == 0):
            dmin = dnear
            dkey = key
        elif (dnear > 0 and dnear < dmin):
            dmin = dnear
            dkey = key
        #print dnear, dmin
        #print 'dkey = '+ dkey
        if (dkey != ''):
            stnum = sn_dics[dkey]
        else: #股本为最小值
            stnum = sn_dics[min(sn_dics, key = sn_dics.get)]
            
    print "find key = "+ dkey +"and  stnum = " + stnum
    return long(stnum)*10000
    
def getMktval(date, sn_dics, df):
    mktval = 0.0
    #get the date close price
    ndate = str(date)
    ndate = ndate[0:4]+ '-' + ndate[4:6]
    #print ndate
    #假定df内对应ndate的只有一行（月）的收盘价
    close = df[df['date'].str.contains(ndate)].close
    if (close.empty): #还没有上市，无收盘价格
        close = 0
    close = float(close)
    #print type(close), close
    #get the date stock num
    stnum = getStocknum(date, sn_dics)
    mktval = close*stnum           
    return mktval
    
    
    
tmpfile = 'd:\\tmp.html'
stock = '603568'#'603605'#'300072' #
#downurl = 'https://www.python.org/events/python-events/'
downurl = 'http://money.finance.sina.com.cn/corp/go.php/vCI_StockStructureHistory/stockid/'+stock+'/stocktype/TotalStock.html'    
#getNetFile(downurl, tmpfile)
sn_dics = parseStnum(downurl)
print sn_dics

df_mprice = ts.get_k_data(stock, ktype='M')
#日期转换成sina格式（对应月份），20171231	20170930	20170630	20170331
#print df_mprice['close']
#fname = "D:\\Programming\\Python\\stockSource\\baobiao\\tmp.xlsx"
fname = "D:\\Document\\PythonDir\\baobiao\\tmp.xlsx"
#getExcelDF(fname)
wb = load_workbook(fname)
s_sheet = wb.get_sheet_by_name('tmp')
rows_len = len(list(s_sheet.rows))
cols_len = len(list(s_sheet.columns))
print rows_len, cols_len
for col in range(1, cols_len+1):
    if col == 1:
        s_sheet.cell(row=rows_len+1, column=col).value = '总市值'
    else:
        #stnum = getStocknum(s_sheet.cell(row=1, column=col).value, sn_dics)
        mktval = getMktval(s_sheet.cell(row=1, column=col).value, sn_dics, df_mprice)
        s_sheet.cell(row=rows_len+1, column=col).value = mktval
wb.save(fname)       